"""
app.py — DocuWizard local web server.

WHY A BROWSER UI AND NOT A HOSTED SITE
--------------------------------------
This binds to 127.0.0.1 (loopback) only. That address is not reachable from any
other machine - not from the network, not from the internet - so "web UI" here
means "a local window that happens to be rendered by the browser you already
have installed". Files are read from and written to this machine only. Nothing
is uploaded anywhere. This keeps the on-device guarantee intact while allowing a
real rule-builder interface, which a plain desktop dialog toolkit cannot
practically provide.

Run it with:   python app.py
It opens http://127.0.0.1:8765 in the default browser.
"""

import json
import os
import queue
import shutil
import threading
import webbrowser

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import (FileResponse, HTMLResponse, JSONResponse,
                               StreamingResponse)
import uvicorn

from core import pipeline, readers, ruleset, sources
from core.engine import RULE_TYPES

import sys

# PyInstaller note: when frozen, read-only resources (static/, samples/) are
# unpacked to a temporary folder (sys._MEIPASS), while anything the user must
# keep - the workspace - has to live NEXT TO THE EXECUTABLE, because the temp
# folder is deleted when the program exits.
if getattr(sys, "frozen", False):
    BUNDLE = sys._MEIPASS                              # read-only resources
    APP_DIR = os.path.dirname(sys.executable)          # writable, persistent
else:
    BUNDLE = APP_DIR = os.path.dirname(os.path.abspath(__file__))

BASE = BUNDLE
STATIC = os.path.join(BUNDLE, "static")
UPLOADS = os.path.join(ruleset.WORKSPACE, "uploads")
SAMPLES = os.path.join(BASE, "samples")

# Uploaded files, generated runs, rulesets and templates all persist in the
# workspace folder between sessions, so previous work can be reopened.
os.makedirs(UPLOADS, exist_ok=True)


app = FastAPI(title="DocuWizard")
ruleset.ensure_starters()

# The most recent run. Output bytes are also buffered in memory so the
# "Download again" buttons keep working even if the run folder is later moved
# or cleaned up by hand.
LAST = {"run": None, "buffers": {}}


def _buffer_outputs(result):
    """Read each generated file into memory so it survives the purge."""
    buffers = {}
    for o in result.get("outputs", []):
        try:
            with open(o["path"], "rb") as fh:
                buffers[o["format"]] = (o["name"], fh.read())
        except OSError:
            continue
    return buffers



@app.get("/", response_class=HTMLResponse)
def index():
    with open(os.path.join(STATIC, "index.html"), encoding="utf-8") as fh:
        return fh.read()


@app.get("/api/meta")
def meta():
    """Everything the UI needs to draw itself."""
    from core.engine import METHODS_BY_EXT, RULE_HELP
    return {"rule_types": RULE_TYPES,
            "methods_by_ext": METHODS_BY_EXT,
            "rule_help": RULE_HELP,
            "supported": readers.SUPPORTED,
            "has_samples": os.path.isdir(SAMPLES) and bool(os.listdir(SAMPLES)),
            "rulesets": ruleset.list_rulesets(),
            "templates": ruleset.list_templates()}


# ---------------------------------------------------------------- rule sets
@app.get("/api/rulesets/{rid}")
def get_ruleset(rid: str):
    data = ruleset.load_ruleset(rid)
    if data is None:
        raise HTTPException(404, "Rule set not found")
    return data


@app.post("/api/rulesets")
async def post_ruleset(payload: dict):
    rid = ruleset.save_ruleset(payload)
    return {"id": rid, "rulesets": ruleset.list_rulesets()}


@app.delete("/api/rulesets/{rid}")
def del_ruleset(rid: str):
    ruleset.delete_ruleset(rid)
    return {"rulesets": ruleset.list_rulesets()}


# ---------------------------------------------------------------- templates
@app.get("/api/templates/{tid}")
def get_template(tid: str):
    data = ruleset.load_template(tid)
    if data is None:
        raise HTTPException(404, "Template not found")
    return data


@app.post("/api/templates")
async def post_template(payload: dict):
    tid = ruleset.save_template(payload)
    return {"id": tid, "templates": ruleset.list_templates()}


@app.delete("/api/templates/{tid}")
def del_template(tid: str):
    ruleset.delete_template(tid)
    return {"templates": ruleset.list_templates()}


# ------------------------------------------------------------------- upload
@app.post("/api/upload")
async def upload(files: list[UploadFile] = File(...)):
    """
    Add files to the local library. Existing files are kept, so previously
    uploaded documents stay available between sessions.
    """
    os.makedirs(UPLOADS, exist_ok=True)
    saved = []
    for f in files:
        name = os.path.basename(f.filename or "unnamed")
        dest = os.path.join(UPLOADS, name)
        with open(dest, "wb") as out:
            shutil.copyfileobj(f.file, out)
        ext = os.path.splitext(name)[1].lower()
        saved.append({"label": name, "size": os.path.getsize(dest),
                      "supported": ext in readers.SUPPORTED,
                      "ext": ext})
    return {"files": saved}


@app.post("/api/samples")
def load_samples():
    """Replace the working files with the bundled cartoon sample set."""
    if not os.path.isdir(SAMPLES):
        raise HTTPException(404, "No sample files are bundled with this copy.")
    os.makedirs(UPLOADS, exist_ok=True)
    saved = []
    for name in sorted(os.listdir(SAMPLES)):
        src = os.path.join(SAMPLES, name)
        if os.path.isfile(src):
            shutil.copy(src, os.path.join(UPLOADS, name))
            ext = os.path.splitext(name)[1].lower()
            saved.append({"label": name, "size": os.path.getsize(src),
                          "supported": ext in readers.SUPPORTED, "ext": ext})
    return {"files": saved}


@app.post("/api/files/clear")
def clear_files():
    """Remove every uploaded file from the local library. Runs are kept."""
    shutil.rmtree(UPLOADS, ignore_errors=True)
    os.makedirs(UPLOADS, exist_ok=True)
    return {"files": []}


@app.delete("/api/files/{label}")
def delete_file(label: str):
    """Remove one file from the local library."""
    path = os.path.join(UPLOADS, os.path.basename(label))
    if os.path.isfile(path):
        os.remove(path)
    return {"files": sources.LocalUpload(UPLOADS).collect()}


def _selected_files(labels):
    """
    Narrow the local library down to the files the user ticked for this run.
    An empty/missing selection means "use everything", so older callers and a
    first-time user both still work.
    """
    everything = sources.LocalUpload(UPLOADS).collect()
    if not labels:
        return everything
    wanted = {str(x) for x in labels}
    return [f for f in everything if f["label"] in wanted]


@app.get("/api/uploaded")
def uploaded():
    return {"files": sources.LocalUpload(UPLOADS).collect()}


@app.get("/api/preview")
def preview(label: str):
    """
    Return a renderable view of one uploaded file so the browser can show it and
    let the user pick anchors by selecting text (or clicking a cell).

    Spreadsheets return a cell grid with real addresses (A1, P10) so clicking a
    cell can fill in the sheet/cell rule directly. Everything else returns the
    ordered blocks the rule engine itself sees, so what the user selects is
    exactly what the rules will match against.
    """
    path = os.path.join(UPLOADS, os.path.basename(label))
    if not os.path.isfile(path):
        raise HTTPException(404, "That file is not uploaded.")

    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        from openpyxl.utils import get_column_letter
        wb = openpyxl.load_workbook(path, data_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for r in range(1, min(ws.max_row or 1, 60) + 1):
                row = []
                for c in range(1, min(ws.max_column or 1, 30) + 1):
                    v = ws.cell(row=r, column=c).value
                    row.append({"ref": f"{get_column_letter(c)}{r}",
                                "v": "" if v is None else str(v)})
                rows.append(row)
            sheets.append({"name": name, "rows": rows})
        return {"kind": "sheet", "sheets": sheets}

    reader, problem = readers.open_reader(path)
    if reader is None:
        raise HTTPException(400, problem)
    blocks = []
    for b in reader.blocks():
        if b["type"] == "paragraph":
            blocks.append({"type": "paragraph", "text": b["text"],
                           "location": b["location"]})
        else:
            blocks.append({"type": "table", "rows": b["rows"],
                           "location": b["location"]})
    return {"kind": "doc", "blocks": blocks}


@app.get("/api/rulesets/{rid}/file")
def download_ruleset(rid: str):
    """Download a rule set as JSON, so it can be shared or version-controlled."""
    path = os.path.join(ruleset.RULESETS_DIR, f"{rid}.json")
    if not os.path.isfile(path):
        raise HTTPException(404, "Rule set not found")
    return FileResponse(path, filename=f"{rid}.ruleset.json",
                        media_type="application/json")


@app.get("/api/templates/{tid}/file")
def download_template(tid: str):
    """Download an output template as JSON."""
    path = os.path.join(ruleset.TEMPLATES_DIR, f"{tid}.json")
    if not os.path.isfile(path):
        raise HTTPException(404, "Template not found")
    return FileResponse(path, filename=f"{tid}.template.json",
                        media_type="application/json")


@app.post("/api/import")
async def import_json(payload: dict):
    """Import a previously downloaded rule set or output template."""
    kind = payload.get("kind")
    data = payload.get("data") or {}
    if kind == "ruleset":
        rid = ruleset.save_ruleset(data)
        return {"id": rid, "rulesets": ruleset.list_rulesets()}
    if kind == "template":
        tid = ruleset.save_template(data)
        return {"id": tid, "templates": ruleset.list_templates()}
    raise HTTPException(400, "Unknown import type.")


# ---------------------------------------------------------------------- run
@app.post("/api/run")
async def run_extraction(payload: dict):
    rs = ruleset.load_ruleset(payload.get("ruleset", ""))
    if rs is None:
        raise HTTPException(400, "Choose a rule set first.")
    tpl = ruleset.load_template(payload.get("template", "")) or {}
    formats = payload.get("formats") or ["docx"]

    files = _selected_files(payload.get("files"))
    if not files:
        raise HTTPException(400, "Choose at least one file first.")

    log_lines = []
    result = pipeline.run(files, rs, tpl, formats, log=log_lines.append)
    LAST["run"] = result
    LAST["buffers"] = _buffer_outputs(result)

    return JSONResponse({
        "ok": True,
        "log": log_lines,
        "warnings": result["warnings"],
        "missing": result["missing"],
        "outputs": [{"format": o["format"], "name": o["name"]}
                    for o in result["outputs"]],
        "fields": [{"name": r["name"], "found": r["found"],
                    "citation": r["citation"], "note": r.get("note"),
                    "preview": _preview(r)} for r in result["records"]],
    })


def _preview(rec):
    """A short, plain-text glimpse of a value for the results table."""
    if not rec["found"]:
        return ""
    v = rec["value"]
    if rec["kind"] == "text":
        s = str(v)
    elif rec["kind"] == "table":
        s = " / ".join(c for c in (v.get("rows") or [[]])[0])
    else:
        parts = []
        for b in v[:3]:
            parts.append(b["text"] if b["type"] == "paragraph"
                         else f"[table {len(b['rows'])} rows]")
        s = " ".join(parts)
    s = " ".join(s.split())
    return s[:160] + ("…" if len(s) > 160 else "")


@app.post("/api/run/stream")
async def run_stream(payload: dict):
    """
    Same as /api/run, but streams progress lines as they happen so the UI can
    show a live log and a progress bar instead of an opaque wait.

    Sends newline-delimited JSON: {"log": "..."} for each step, then a final
    {"result": {...}} object.
    """
    rs = ruleset.load_ruleset(payload.get("ruleset", ""))
    if rs is None:
        raise HTTPException(400, "Choose a ruleset first.")
    tpl = ruleset.load_template(payload.get("template", "")) or {}
    formats = payload.get("formats") or ["docx"]
    files = _selected_files(payload.get("files"))
    if not files:
        raise HTTPException(400, "Choose at least one file first.")

    q = queue.Queue()
    DONE = object()

    def worker():
        try:
            result = pipeline.run(files, rs, tpl, formats, log=q.put)
            LAST["run"] = result
            LAST["buffers"] = _buffer_outputs(result)
            q.put(("__RESULT__", result))
        except Exception as exc:
            q.put(("__ERROR__", str(exc)))
        finally:
            q.put(DONE)

    threading.Thread(target=worker, daemon=True).start()

    # Total steps = files read + rules applied + formats written, used for the
    # progress bar. Approximate is fine; the bar is reassurance, not telemetry.
    total = len(files) + len(rs.get("fields", [])) + len(formats) + 1

    def stream():
        seen = 0
        while True:
            item = q.get()
            if item is DONE:
                break
            if isinstance(item, tuple) and item[0] == "__RESULT__":
                r = item[1]
                yield json.dumps({"result": {
                    "ok": True,
                    "log": [],
                    "warnings": r["warnings"],
                    "missing": r["missing"],
                    "outputs": [{"format": o["format"], "name": o["name"]}
                                for o in r["outputs"]],
                    "fields": [{"name": x["name"], "found": x["found"],
                                "citation": x["citation"], "note": x.get("note"),
                                "preview": _preview(x)} for x in r["records"]],
                }}) + "\n"
            elif isinstance(item, tuple) and item[0] == "__ERROR__":
                yield json.dumps({"error": item[1]}) + "\n"
            else:
                seen += 1
                yield json.dumps({"log": str(item),
                                  "pct": min(96, int(seen * 100 / max(total, 1)))}) + "\n"

    return StreamingResponse(stream(), media_type="application/x-ndjson")


@app.get("/api/download/{fmt}")
def download(fmt: str):
    """
    Serve the last run's output from memory. The file no longer exists on disk -
    it is purged as soon as the run completes - so re-downloading reads the
    buffered bytes instead.
    """
    from fastapi.responses import Response
    buf = LAST.get("buffers", {}).get(fmt)
    if not buf:
        raise HTTPException(404, "Nothing has been extracted yet.")
    name, data = buf
    return Response(content=data, media_type="application/octet-stream",
                    headers={"Content-Disposition": f'attachment; filename="{name}"'})


def main():
    url = "http://127.0.0.1:8765"
    threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    print(f"DocuWizard running at {url}  (local only — press Ctrl+C to stop)")
    uvicorn.run(app, host="127.0.0.1", port=8765, log_level="warning")


if __name__ == "__main__":
    main()
